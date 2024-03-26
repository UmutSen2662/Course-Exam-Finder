from glob import glob
import tkinter.ttk as ttk
import tkinter as tk
import openpyxl, csv, os

#Get absolute path to resource, works for dev and for PyInstaller
def resource_path(relative_path):
    base_path = os.getcwd()
    return glob(base_path + relative_path)


def read_create(file_path, classes):
    with open(file_path, "r", encoding="utf-8") as file:
        csv_reader = csv.reader(file)

        workBook = openpyxl.Workbook()
        sheet = workBook.active
        sheet.column_dimensions['A'].width = 40.0
        sheet.column_dimensions['B'].width = 50.0
        sheet.column_dimensions['C'].width = 80.0
        sheet.append(("Course Exam", "Start Time", "Classrooms"))
        for cell in sheet[1]:
            cell.fill = openpyxl.styles.PatternFill(start_color="afafaf", fill_type="solid")

        if "Final" in file_path:
            for row in csv_reader:
                if row[0][:8].replace(" ","").upper() in classes.replace(" ","").upper():
                    sheet.append((row[0], row[1] + " " + row[2][:5], row[3]))
        else:
            for row in csv_reader:
                if row[0][:8].replace(" ","").upper() in classes.replace(" ","").upper():
                    sheet.append((row[0], row[1], row[3]))

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = openpyxl.styles.Font(size=16)
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='medium', color='000000'),
                                                     right=openpyxl.styles.Side(border_style='medium', color='000000'),
                                                     top=openpyxl.styles.Side(border_style='medium', color='000000'),
                                                     bottom=openpyxl.styles.Side(border_style='medium', color='000000'))

        if "FINAL" in file_path.upper():
            workBook.save('Finals.xlsx')
        elif "MIDTERM" in file_path.upper():
            workBook.save('Midterms.xlsx')
        else:
            workBook.save('Exams.xlsx')


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.configure(bg="#222222")
        self.title("Course Exam Finder")
        self.minsize(self.winfo_screenwidth()//3, self.winfo_screenheight()//5)
        style = ttk.Style(self)
        style.configure("TButton", font=(None, 16), background="#222222")

        label1 = ttk.Label(self, text="Course Exam Finder", foreground="#eeeeee", background="#222222", font=(None, 20))
        label1.pack(pady=10)

        self.entry1 = tk.Entry(self, font=(None, 18), fg="#eeeeee", insertbackground="#eeeeee", background="#2f2f2f")
        self.entry1.pack(padx=15, ipady=3, fill="x")
        self.entry1.insert(0, "Enter course codes")
        self.entry1.bind("<FocusIn>", lambda event: self.delet_entry(self.entry1))
        self.entry1.bind("<FocusOut>", lambda event: self.put_placeholder(self.entry1, "Enter course codes"))

        self.entry2 = tk.Entry(self, font=(None, 18), fg="#eeeeee", insertbackground="#eeeeee", background="#2f2f2f")
        self.entry2.pack(padx=15, ipady=3, fill="x", pady=15)
        self.entry2.insert(0, "Enter CSV file name")
        self.entry2.bind("<FocusIn>", lambda event: self.delet_entry(self.entry2))
        self.entry2.bind("<FocusOut>", lambda event: self.put_placeholder(self.entry2, "Enter CSV file name"))

        button = ttk.Button(self, text="Find", width=6, command=self.create_exam_file)
        button.pack(padx=20, ipady=5)
        self.bind("<Return>", lambda event: self.create_exam_file())

    def delet_entry(self, entry):
        if entry.get() == "Enter course codes" or entry.get() == "Enter CSV file name" or entry.get() == "No CSV file found":
            entry.delete(0, tk.END)

    def put_placeholder(self, entry, text):
        if entry.get() == "":
            entry.insert(0, text)

    def create_exam_file(self):
        course_codes = self.entry1.get()
        csv_name = self.entry2.get()
        file_path = resource_path("\*"+csv_name+"*.csv")
        if not file_path:
            self.entry2.delete(0, tk.END)
            self.entry2.insert(0, "No CSV file found")
        else:
            read_create(file_path[0], course_codes)
            app.destroy()

app = App()
app.mainloop()
