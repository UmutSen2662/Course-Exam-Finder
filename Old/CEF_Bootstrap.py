from ttkbootstrap.constants import *
import ttkbootstrap as ttk
from glob import glob
import openpyxl
import csv
import os

#Get absolute path to resource, works for dev and for PyInstaller
def resource_path(relative_path):
    try:
        base_path = os.path.abspath(".") + "\\"
        file_path = glob(base_path + relative_path)
        return file_path
    except Exception:
        base_path = os.path.abspath(".") + "\\CEF\\"
        file_path = glob(base_path + relative_path)
        return file_path

def read_create(file_path, classes):
    with open(file_path) as file:
        csv_reader = csv.reader(file)

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 40.0
        sheet.column_dimensions['B'].width = 50.0
        sheet.column_dimensions['C'].width = 80.0
        sheet.append(("Course Exam", "Start Time", "Classrooms"))
        for cell in sheet[1]:
            cell.fill = openpyxl.styles.PatternFill(start_color="afafaf", fill_type="solid")

        if "Final" in file_path:
            for row in csv_reader:
                if row[0][:8].replace(" ","").upper() in classes.replace(" ","").upper():
                    sheet.append((row[0], row[1] + " " + row[2][:5], row[3]))
        else:
            for row in csv_reader:
                if row[0][:8].replace(" ","").upper() in classes.replace(" ","").upper():
                    sheet.append((row[0], row[1], row[3]))

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = openpyxl.styles.Font(size=16)
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='medium', color='000000'),
                                                    right=openpyxl.styles.Side(border_style='medium', color='000000'),
                                                    top=openpyxl.styles.Side(border_style='medium', color='000000'),
                                                    bottom=openpyxl.styles.Side(border_style='medium', color='000000'))

        if "FINAL" in file_path.upper():
            wb.save('Finals.xlsx')
        elif "MIDTERM" in file_path.upper():
            wb.save('Midterms.xlsx')
        else:
            wb.save('Exams.xlsx')

class App(ttk.Window):
    def __init__(self):
        super().__init__()
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        self.title("Course Exam Finder")
        self.minsize(screen_width//3, screen_height//5)
        style = ttk.Style("darkly")
        style.configure("TButton", font=(None, 14))

        label1 = ttk.Label(self, text="Course Exam Finder", foreground="white", font=(None, 20))
        label1.pack()

        self.entry1 = ttk.Entry(self, bootstyle=(LIGHT), font=(None, 18))
        self.entry1.pack(padx=15, ipady=2, fill="x", pady=15)
        self.entry1.insert(0, "Enter course codes")
        self.entry1.bind("<FocusIn>", lambda event: self.delet_e1())

        self.entry2 = ttk.Entry(self, bootstyle=(LIGHT), font=(None, 18))
        self.entry2.pack(padx=15, ipady=2, fill="x")
        self.entry2.insert(0, "Enter CSV file name")
        self.entry2.bind("<FocusIn>", lambda event: self.delet_e2())

        button = ttk.Button(self, text="Find", width=6, bootstyle=INFO, command=self.do_the_thing)
        button.pack(padx=20, ipady=3, pady=15)
        self.bind("<Return>", lambda event: self.do_the_thing())
    
    def delet_e1(self):
        if self.entry1.get() == "Enter course codes":
            self.entry1.delete(0, END)

    def delet_e2(self):
        if self.entry2.get() == "No CSV file found" or self.entry2.get() == "Enter CSV file name":
            self.entry2.delete(0, END)

    def do_the_thing(self):
        courses = self.entry1.get()
        csv_name = self.entry2.get()
        file_path = resource_path("*"+csv_name+"*.csv")

        if not file_path:
            self.entry2.delete(0, END)
            self.entry2.insert(0, "No CSV file found")
        else:
            read_create(file_path[0], courses)
            app.destroy()


app = App()
app.mainloop()
